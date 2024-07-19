import openpyxl

wb50 = openpyxl.load_workbook('20250.xlsx')
sheet50 = wb50.active

wb90 = openpyxl.load_workbook('20290.xlsx')
sheet90 = wb90.active

wb91 = openpyxl.load_workbook('20291.xlsx')
sheet91 = wb91.active

search_sign = input("Введите предмет поиска (День,Время,Неделя,Аудитория,Предмет,Преподаватель) ")

row = ["A","B","C","D","E","F","G"]

for j in range(65,72):
    for i in range(1,29):
        if search_sign == sheet50[f'{chr(j)}{i}'].value:
            for b in range (0,7):
                print(str(sheet50[f'{row[b]}{i}'].value),end = " ")
            print("")
        if search_sign == sheet90[f'{chr(j)}{i}'].value:
            for c in range(0, 7):
                print(str(sheet90[f'{row[c]}{i}'].value), end=" ")
            print("")
        if search_sign == sheet91[f'{chr(j)}{i}'].value:
            for g in range(0, 7):
                print(str(sheet91[f'{row[g]}{i}'].value), end=" ")
            print("")