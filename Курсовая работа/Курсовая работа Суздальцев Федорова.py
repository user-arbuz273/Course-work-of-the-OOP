import openpyxl
import pandas as pd

class Choose():
    def __init__(self):
        self.cols = [0, 1, 2, 3, 4, 5, 6]
        self.groups = ['20250', '20290', '20291']

    def chosen_group_viewing(self, chosen_group):
        if chosen_group in self.groups:
            df = pd.read_excel(f'{chosen_group}.xlsx', usecols=self.cols)
            pd.set_option('display.max_columns', None,
                          'display.max_rows', None)
            pd.options.display.expand_frame_repr = False
            print(df)

            print("Хотите посмотреть расписание других групп?")
            choice1 = input("Да/Нет\n")
            if choice1 == "Да":
                return self.chosen_group_viewing(input('Введите номер группы: '))
            else:
                start()
        else:
            print('Введён неверный номер группы номер группы. Повторите попытку: ')
            return self.chosen_group_viewing(input())

# ch = Choose()
# ch.chosen_group_viewing(input())

class Sorting():
    def __init__(self):
        self.row = ["A", "B", "C", "D", "E", "F", "G"]

    def sorted_schedule(self):
        wb50 = openpyxl.load_workbook('20250.xlsx')
        sheet50 = wb50.active
        wb90 = openpyxl.load_workbook('20290.xlsx')
        sheet90 = wb90.active
        wb91 = openpyxl.load_workbook('20291.xlsx')
        sheet91 = wb91.active
        search_sign = input("Введите предмет поиска (День,Время,Неделя,Аудитория,Предмет,Преподаватель) ")
        for j in range(65, 72):
            for i in range(1, 29):
                if search_sign == sheet50[f'{chr(j)}{i}'].value:
                    for b in range(0, 7):
                        print(str(sheet50[f'{self.row[b]}{i}'].value), end=" ")
                    print("")
                if search_sign == sheet90[f'{chr(j)}{i}'].value:
                    for c in range(0, 7):
                        print(str(sheet90[f'{self.row[c]}{i}'].value), end=" ")
                    print("")
                if search_sign == sheet91[f'{chr(j)}{i}'].value:
                    for g in range(0, 7):
                        print(str(sheet91[f'{self.row[g]}{i}'].value), end=" ")
                    print("")
        choice2=input('Повторить поиск? (Да/Нет)\n')
        if choice2 == "Да":
            self.sorted_schedule()
        else:
            start()

# sr = Sorting()
# sr.sorted_schedule()

class Creation():
    def __init__(self):
        self.name = ['0', 'День недели', 'Время', 'Неделя', 'Аудитория', 'Группа', 'Предмет', 'Преподаватель']
        self.name2 = ['0', 'a', 'b', 'c', 'd', 'e', 'f', 'g']
        self.name3 = ['0', 'A', 'B', 'C', 'D', 'E', 'F', 'G']
    def creating_schedule(self):
        wb = openpyxl.Workbook()
        list = wb.active
        namefile=input('Введите название файла: ')
        for num in range(1, 8):
            f'{self.name2}1' == list[f"{self.name3[num]}1"]
            list[f"{self.name3[num]}1"] = f'{self.name[num]}'
        for i in range(2, 31):
            m = input('Добавить строчку? (Да/Нет) ')
            if m == 'Да':
                for j in range(0, 7):
                    n = input(f'Введите {self.name[j+1]} ')
                    a2 = list[f"{self.name3[j + 1]}{i}"]
                    list[f'{self.name3[j + 1]}{i}'] = n
            else:
                break
        wb.save(f'{namefile}.xlsx')
        choice3 = input('Создать еще одно расписание? (Да/Нет)\n')
        if choice3 == "Да":
            self.creating_schedule()
        else:
            start()

# cr= Creation()
# cr.creating_schedule()

def start():
    print("Доброго времени суток ! \nЭто приложения позволяет:")
    func = ['Найти расписание для учебных групп','Найти расписание по отдельности','Создать свое расписание']
    for i in range(3):
        print(f'{i+1}',f'{func[i]}')
    st=input('Выберите, что вы хотите сделать: \nЕсли вы хотите выйти из программы, напишите "Выйти"\n')
    if st == '1':
        ch = Choose()
        ch.chosen_group_viewing(input('Введите номер группы: '))
    if st == '2':
        sr = Sorting()
        sr.sorted_schedule()
    if st == '3':
        cr= Creation()
        cr.creating_schedule()
    if st == 'Выйти':
        return

start()