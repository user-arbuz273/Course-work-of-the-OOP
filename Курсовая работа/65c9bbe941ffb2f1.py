import pandas as pd
import openpyxl

class Choose():
    def __init__(self):
        self.cols = [0, 1, 2, 3, 4, 5, 6]
        self.groups = ['20250', '20290', '20291']

        self.days = ["A", "B", "C", "D", "E", "F", "G"]
        self.wb50 = openpyxl.load_workbook('20250.xlsx')
        self.sheet50 = self.wb50.active
        self.wb90 = openpyxl.load_workbook('20290.xlsx')
        self.sheet90 = self.wb90.active
        self.wb91 = openpyxl.load_workbook('20291.xlsx')
        self.sheet91 = self.wb91.active

    def chosen_group_viewing(self, chosen_group):
        if chosen_group in self.groups:
            df = pd.read_excel(f'{chosen_group}.xlsx', usecols=self.cols)
            pd.set_option('display.max_columns', None,
                          'display.max_rows', None)
            pd.options.display.expand_frame_repr = False
            print(df)

            print("Хотите посмотреть расписание других групп?")
            choice1 = input("Да/Нет\n")
            if choice1 == "Да":
                return self.chosen_group_viewing(input("Введите группу: "))
        else:
            print('Введён неверный номер группы номер группы. Повторите попытку: ')
            return self.chosen_group_viewing(input())

    def search_engine(self):

def welcome():
    print("Доброго времени суток ! \nЭто приложения позволяет: \n1.Найти расписание для учебных групп \n2.Создать свое расписание \n3.Поиск расписание по отдельности \nЧтобы вы хотели выбрать ? (1,2,3) ")
    choice_prog = input()
    if choice_prog == "1":
        ch.chosen_group_viewing(input("Введите группу: "))
    if choice_prog == "2":
        ch.search_engine()
    if choice_prog == "3":

ch = Choose()
welcome()









