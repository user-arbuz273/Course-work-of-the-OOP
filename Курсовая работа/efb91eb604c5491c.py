import openpyxl

days = ["A","B","C","D","E","F","G"]

wb50 = openpyxl.load_workbook('20250.xlsx')
sheet50 = wb50.active

wb90 = openpyxl.load_workbook('20290.xlsx')
sheet90 = wb90.active

wb91 = openpyxl.load_workbook('20291.xlsx')
sheet91 = wb91.active



search_sign = input("Введите предмет поиска (День,Время,Неделя,Аудитория,Предмет,Преподаватель) ")

if search_sign == "День":

    # День
    search_object = input("Введите объект поиска ")

    for j in range(1,29):
        if search_object == sheet50[f'A{j}'].value:
            for i in range (0,7):
                print(sheet50[f'{days[i]}{j}'].value, end=" ")
            print("")
    print("")
    for j in range(1,29):
        if search_object == sheet90[f'A{j}'].value:
            for i in range (0,7):
                print(sheet90[f'{days[i]}{j}'].value, end=" ")
            print("")
    print("")
    for j in range(1,30):
        if search_object == sheet91[f'A{j}'].value:
            for i in range (0,7):
                print(sheet91[f'{days[i]}{j}'].value, end=" ")
            print("")

if search_sign == "Время":

    # Время
    search_object = input("Введите объект поиска ")

    for j in range(1, 29):
        if search_object == sheet50[f'B{j}'].value:
            for i in range(0, 7):
                print(sheet50[f'{days[i]}{j}'].value, end=" ")
            print("")
    print("")
    for j in range(1, 29):
        if search_object == sheet90[f'B{j}'].value:
            for i in range(0, 7):
                print(sheet90[f'{days[i]}{j}'].value, end=" ")
            print("")
    print("")
    for j in range(1, 30):
        if search_object == sheet91[f'B{j}'].value:
            for i in range(0, 7):
                print(sheet91[f'{days[i]}{j}'].value, end=" ")
            print("")

if search_sign == "Неделя":

    # Неделя
    search_object = input("Введите объект поиска ")

    for j in range(1, 29):
        if search_object == sheet50[f'C{j}'].value:
            for i in range(0, 7):
                print(sheet50[f'{days[i]}{j}'].value, end=" ")
            print("")
    print("")
    for j in range(1, 29):
        if search_object == sheet90[f'C{j}'].value:
            for i in range(0, 7):
                print(sheet90[f'{days[i]}{j}'].value, end=" ")
            print("")
    print("")
    for j in range(1, 30):
        if search_object == sheet91[f'C{j}'].value:
            for i in range(0, 7):
                print(sheet91[f'{days[i]}{j}'].value, end=" ")
            print("")


if search_sign == "Аудитория":

    # Аудитория
    search_object = input("Введите объект поиска ")

    for j in range(1, 29):
        if search_object == sheet50[f'D{j}'].value:
            for i in range(0, 7):
                print(sheet50[f'{days[i]}{j}'].value, end=" ")
            print("")
    print("")
    for j in range(1, 29):
        if search_object == sheet90[f'D{j}'].value:
            for i in range(0, 7):
                print(sheet90[f'{days[i]}{j}'].value, end=" ")
            print("")
    print("")
    for j in range(1, 30):
        if search_object == sheet91[f'D{j}'].value:
            for i in range(0, 7):
                print(sheet91[f'{days[i]}{j}'].value, end=" ")
            print("")
            

if search_sign == "Предмет":

    # Предмет
    search_object = input("Введите объект поиска ")

    for j in range(1, 29):
        if search_object == sheet50[f'F{j}'].value:
            for i in range(0, 7):
                print(sheet50[f'{days[i]}{j}'].value, end=" ")
            print("")
    print("")
    for j in range(1, 29):
        if search_object == sheet90[f'F{j}'].value:
            for i in range(0, 7):
                print(sheet90[f'{days[i]}{j}'].value, end=" ")
            print("")
    print("")
    for j in range(1, 30):
        if search_object == sheet91[f'F{j}'].value:
            for i in range(0, 7):
                print(sheet91[f'{days[i]}{j}'].value, end=" ")
            print("")


if search_sign == "Преподаватель":

    # Преподаватель
    search_object = input("Введите объект поиска ")

    for j in range(1, 29):
        if search_object == sheet50[f'G{j}'].value:
            for i in range(0, 7):
                print(sheet50[f'{days[i]}{j}'].value, end=" ")
            print("")
    print("")
    for j in range(1, 29):
        if search_object == sheet90[f'G{j}'].value:
            for i in range(0, 7):
                print(sheet90[f'{days[i]}{j}'].value, end=" ")
            print("")
    print("")
    for j in range(1, 30):
        if search_object == sheet91[f'G{j}'].value:
            for i in range(0, 7):
                print(sheet91[f'{days[i]}{j}'].value, end=" ")
            print("")
